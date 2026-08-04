"""Cross-surface regression tests for the parametric-label invariants.

T-01 (audit finding B-01) — descriptive surfaces must agree with the
PDF on every parametric value:

  * The XLSX Device List and the UI device-breakdown rows never ship a
    literal template token (``XXX`` / ``XX``) from SIGN_DESCRIPTIONS.
  * On reduced-speed plans the two R2-1 faces (work-zone entrance
    posting vs downstream restoration) get separate rows with the
    actual limits, instead of one merged "SPEED LIMIT XX" line.

T-02 (audit finding B-02) — the audit ``advance.sign_table`` covers
exactly the upstream SIGN_GENERIC placements (mirror pairs deduped):
every sign the layout ships upstream of the taper has a row, and no
phantom rows exist.  Set equality both ways, so a future hand-added or
hand-dropped row fails loudly.

All four S-630-1 fixtures (Cases 11 / 11b / 26 / 27) run through both
invariants — the same bodies the Phase 5 harness pins.
"""

from __future__ import annotations

import re
from typing import Any

import pytest
from openpyxl import load_workbook

from src.api.render_api import _build_device_breakdown
from src.api.schemas import FlaggerLaneClosureScenario, ShoulderScenario, scenario_to_call
from src.export.device_list import export_device_list
from src.export.quote_generator import generate_quote
from src.narrative.crew_narrative import build_narrative_context
from src.rendering.plan_sheet import _scenario_label
from src.rules.devices import DeviceType
from src.rules.validators import (
    DevicePlacement,
    ScenarioParams,
    scenario_display_name,
    scenario_display_name_short,
)

from ._harness import (
    CASE_11_GENERAL_BODY,
    CASE_11B_BODY,
    CASE_26_BODY,
    CASE_27_BODY,
    placements_and_audit,
)

BODIES: dict[str, dict[str, Any]] = {
    "case_11": CASE_11_GENERAL_BODY,
    "case_11b": CASE_11B_BODY,
    "case_26": CASE_26_BODY,
    "case_27": CASE_27_BODY,
}

# (posted, work-zone) speeds for the reduced-speed fixtures — used to
# assert the R2-1 faces carry the right limits.
REDUCED_SPEEDS: dict[str, tuple[int, int]] = {
    "case_11b": (55, 50),
    "case_26": (65, 60),
    "case_27": (75, 65),
}

# A literal XX / XXX placeholder surviving into a rendered description.
_TEMPLATE_TOKEN = re.compile(r"\bX{2,3}\b")


def _pipeline(body: dict[str, Any]) -> tuple[list[DevicePlacement], ScenarioParams]:
    scenario = ShoulderScenario.model_validate(body)
    params, generator, kwargs = scenario_to_call(scenario)
    return generator(params, **kwargs), params


def _xlsx_device_rows(placements, params, tmp_path) -> list[tuple]:
    path = tmp_path / "devices.xlsx"
    export_device_list(placements, params, str(path))
    wb = load_workbook(str(path), read_only=True)
    rows = list(wb["Device List"].iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# T-01 — XLSX
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("name", sorted(BODIES))
def test_xlsx_descriptions_carry_no_template_tokens(name: str, tmp_path) -> None:
    """No Device List description ships a literal XX/XXX placeholder."""
    placements, params = _pipeline(BODIES[name])
    for row in _xlsx_device_rows(placements, params, tmp_path):
        description = str(row[2])
        assert not _TEMPLATE_TOKEN.search(description), (
            f"{name}: XLSX description carries a literal template token: {description!r}"
        )


@pytest.mark.parametrize("name", sorted(REDUCED_SPEEDS))
def test_xlsx_splits_r2_1_faces_on_reduced_plans(name: str, tmp_path) -> None:
    """Reduced-speed plans get two R2-1 rows (entrance + restoration),
    qty 2 each (mirrored on the divided fixtures), with the actual
    limits — not one merged row of 4."""
    posted, wz_speed = REDUCED_SPEEDS[name]
    placements, params = _pipeline(BODIES[name])
    rows = _xlsx_device_rows(placements, params, tmp_path)
    r2_1_rows = [r for r in rows if str(r[2]).startswith("R2-1 ")]
    assert len(r2_1_rows) == 2, f"{name}: expected 2 R2-1 rows, got {r2_1_rows}"
    descriptions = sorted(str(r[2]) for r in r2_1_rows)
    quantities = [r[5] for r in r2_1_rows]
    assert quantities == [2, 2], f"{name}: expected qty 2 per face, got {quantities}"
    assert descriptions[0] == (f"R2-1 SPEED LIMIT {wz_speed} (work-zone speed posting)"), (
        descriptions
    )
    assert descriptions[1] == (f"R2-1 SPEED LIMIT {posted} (posted-speed restoration)"), (
        descriptions
    )


# ---------------------------------------------------------------------------
# T-01 — UI device breakdown
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("name", sorted(BODIES))
def test_breakdown_descriptions_carry_no_template_tokens(name: str) -> None:
    """No device-breakdown row ships a literal XX/XXX placeholder."""
    placements, params = _pipeline(BODIES[name])
    for row in _build_device_breakdown(placements, params):
        assert not _TEMPLATE_TOKEN.search(str(row["device"])), (
            f"{name}: breakdown device description carries a literal template token: {row!r}"
        )


@pytest.mark.parametrize("name", sorted(REDUCED_SPEEDS))
def test_breakdown_splits_r2_1_faces_on_reduced_plans(name: str) -> None:
    """Same R2-1 face split on the device-breakdown panel."""
    posted, wz_speed = REDUCED_SPEEDS[name]
    placements, params = _pipeline(BODIES[name])
    rows = [r for r in _build_device_breakdown(placements, params) if r["code"] == "R2-1"]
    assert len(rows) == 2, f"{name}: expected 2 R2-1 rows, got {rows}"
    devices = sorted(str(r["device"]) for r in rows)
    assert all(r["qty"] == 2 for r in rows), rows
    assert devices[0] == f"SPEED LIMIT {wz_speed} (work-zone speed posting)", devices
    assert devices[1] == f"SPEED LIMIT {posted} (posted-speed restoration)", devices


# ---------------------------------------------------------------------------
# T-01 — crew narrative Required Equipment list (UX-10)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("name", sorted(BODIES))
def test_narrative_equipment_carries_no_template_tokens(name: str) -> None:
    """The Required Equipment bullets must resolve parametric sign
    legends (e.g. ``G20-1 ROAD CONSTRUCTION (NEXT XXX FT)``) to real
    values instead of leaking the bare template token — the same
    invariant the XLSX device list and UI breakdown already hold, now
    extended to the narrative equipment surface (UX-10)."""
    placements, params = _pipeline(BODIES[name])
    bullets = build_narrative_context(placements, params)["equipment_bullets"]
    assert not _TEMPLATE_TOKEN.search(bullets), (
        f"{name}: narrative equipment list carries a literal template token:\n{bullets}"
    )


# ---------------------------------------------------------------------------
# T-01 — quote Equipment Detail (T4-1 closure, Refs #101)
#
# The quote must route through the same shared helpers as the XLSX /
# breakdown / narrative (substitute_sign_description, schedule_key,
# device_row_sort_key, cone_display_name) — asserted here on the
# generated workbook content, not on rate math.
# ---------------------------------------------------------------------------

# Low-speed variant of the Case 11 body: 40 mph on an undivided rural
# road, below the MUTCD §6F.65 36-inch cone threshold — exercises the
# 28-inch branch of cone_display_name on the export surfaces.
CASE_11_LOW_SPEED_BODY: dict[str, Any] = {
    **CASE_11_GENERAL_BODY,
    "roadType": "rural_undivided",
    "speed": 40,
    "divided": False,
}


def _quote_equipment_rows(placements, params, tmp_path) -> list[tuple]:
    """Data rows of the quote's Equipment Detail sheet (skips SUBTOTAL)."""
    path = tmp_path / "quote.xlsx"
    generate_quote(placements, params, output_path=str(path))
    wb = load_workbook(str(path), read_only=True)
    rows = [
        r
        for r in wb["Equipment Detail"].iter_rows(min_row=2, values_only=True)
        if isinstance(r[0], int)
    ]
    wb.close()
    return rows


@pytest.mark.parametrize("name", sorted(BODIES))
def test_quote_descriptions_carry_no_template_tokens(name: str, tmp_path) -> None:
    """No Equipment Detail description ships a literal XX/XXX placeholder,
    and no labeled sign falls back to its bare code (the W3-5(NN)
    dict-miss case)."""
    placements, params = _pipeline(BODIES[name])
    rows = _quote_equipment_rows(placements, params, tmp_path)
    assert rows, f"{name}: quote Equipment Detail has no data rows"
    for row in rows:
        description = str(row[3])
        assert not _TEMPLATE_TOKEN.search(description), (
            f"{name}: quote description carries a literal template token: {description!r}"
        )
        if row[1] == "SIGN_GENERIC" and row[2]:
            assert description != str(row[2]), (
                f"{name}: labeled sign fell back to its bare code: {description!r}"
            )


@pytest.mark.parametrize("name", sorted(REDUCED_SPEEDS))
def test_quote_splits_r2_1_faces_on_reduced_plans(name: str, tmp_path) -> None:
    """Same R2-1 face split the XLSX/breakdown already hold: two rows
    (entrance + restoration), qty 2 each, with the actual limits."""
    posted, wz_speed = REDUCED_SPEEDS[name]
    placements, params = _pipeline(BODIES[name])
    rows = _quote_equipment_rows(placements, params, tmp_path)
    r2_1_rows = [r for r in rows if str(r[3]).startswith("R2-1 ")]
    assert len(r2_1_rows) == 2, f"{name}: expected 2 R2-1 rows, got {r2_1_rows}"
    descriptions = sorted(str(r[3]) for r in r2_1_rows)
    quantities = [r[4] for r in r2_1_rows]
    assert quantities == [2, 2], f"{name}: expected qty 2 per face, got {quantities}"
    assert descriptions[0] == (f"R2-1 SPEED LIMIT {wz_speed} (work-zone speed posting)"), (
        descriptions
    )
    assert descriptions[1] == (f"R2-1 SPEED LIMIT {posted} (posted-speed restoration)"), (
        descriptions
    )


@pytest.mark.parametrize("name", sorted(BODIES))
def test_quote_rows_match_xlsx_order_and_descriptions(name: str, tmp_path) -> None:
    """Quote Equipment Detail agrees with the XLSX Device List row for
    row — same device_row_sort_key order, byte-identical descriptions,
    identical quantities, signs leading."""
    placements, params = _pipeline(BODIES[name])
    quote_rows = _quote_equipment_rows(placements, params, tmp_path)
    xlsx_rows = _xlsx_device_rows(placements, params, tmp_path)
    assert [str(r[3]) for r in quote_rows] == [str(r[2]) for r in xlsx_rows], (
        f"{name}: quote and XLSX descriptions disagree (content or order)"
    )
    assert [r[4] for r in quote_rows] == [r[5] for r in xlsx_rows], (
        f"{name}: quote and XLSX quantities disagree"
    )
    assert quote_rows[0][1] == "SIGN_GENERIC", (
        f"{name}: quote does not lead with signs: first row {quote_rows[0]!r}"
    )


def test_quote_sign_descriptions_appear_in_narrative(tmp_path) -> None:
    """Every quote sign description also appears verbatim in the crew
    narrative's Required Equipment bullets (shared substitution helper
    ⇒ byte-identity across the three surfaces)."""
    placements, params = _pipeline(CASE_26_BODY)
    quote_rows = _quote_equipment_rows(placements, params, tmp_path)
    bullets = build_narrative_context(placements, params)["equipment_bullets"]
    sign_descriptions = [str(r[3]) for r in quote_rows if r[1] == "SIGN_GENERIC" and r[2]]
    assert sign_descriptions, "no labeled sign rows in the quote"
    for description in sign_descriptions:
        assert description in bullets, (
            f"quote sign description not found in narrative equipment bullets: {description!r}"
        )


@pytest.mark.parametrize(
    ("body", "expected"),
    [
        (CASE_11_GENERAL_BODY, "Traffic Cone (36-inch)"),
        (CASE_11_LOW_SPEED_BODY, "Traffic Cone (28-inch)"),
    ],
    ids=["55mph_36in", "40mph_28in"],
)
def test_cone_description_resolves_size_on_quote_and_xlsx(
    body: dict[str, Any], expected: str, tmp_path
) -> None:
    """Quote and XLSX resolve the cone size via cone_display_name(speed),
    matching the narrative/UI/plan-sheet legend text."""
    placements, params = _pipeline(body)
    quote_cones = [r for r in _quote_equipment_rows(placements, params, tmp_path) if r[1] == "CONE"]
    xlsx_cones = [r for r in _xlsx_device_rows(placements, params, tmp_path) if r[1] == "CONE"]
    assert quote_cones and xlsx_cones, "no cone row on one of the surfaces"
    assert str(quote_cones[0][3]) == expected
    assert str(xlsx_cones[0][2]) == expected
    bullets = build_narrative_context(placements, params)["equipment_bullets"]
    assert expected in bullets


# ---------------------------------------------------------------------------
# T-02 — audit sign_table covers the upstream placements exactly
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("name", sorted(BODIES))
def test_audit_sign_table_matches_upstream_placements(name: str) -> None:
    """Every upstream SIGN_GENERIC placement (mirror-deduped) has a
    sign_table row, and the table has no rows without a placement."""
    placements, audit = placements_and_audit(BODIES[name])
    taper_start = (
        audit["geometry_validation"]["work_zone_ft"]
        + audit["buffer"]["buffer_ft"]
        + audit["taper"]["L_required_ft"]
    )
    expected = {
        (p.label, round(p.station_ft))
        for p in placements
        if p.device_type == DeviceType.SIGN_GENERIC and p.label and p.station_ft > taper_start
    }
    actual = {
        (row["Code"], round(float(row["Station (ft)"].replace(",", ""))))
        for row in audit["advance"]["sign_table"]
    }
    assert actual == expected, (
        f"{name}: audit sign_table disagrees with upstream placements.\n"
        f"  missing from table: {sorted(expected - actual)}\n"
        f"  phantom rows:       {sorted(actual - expected)}"
    )


# ---------------------------------------------------------------------------
# T-03 — closure-type label reads the real lane count (Refs #118)
#
# scenario_display_name derives the undivided lane claim from
# ``num_lanes`` (total-lane naming, so num_lanes=1 → "2-Lane") and the
# single-sourced string must reach every full-label surface
# byte-identically: plan-sheet title block, XLSX Summary, quote header,
# crew narrative.  The PARAMETERS box short form never carries the
# count — pinned here so a truncation change is a loud failure.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("lanes", [1, 2, 3, 4])
def test_undivided_label_reads_lane_count_on_every_surface(lanes: int, tmp_path) -> None:
    body = {
        **CASE_11_GENERAL_BODY,
        "roadType": "rural_undivided",
        "divided": False,
        "lanes": lanes,
        # 10-ft lanes keep lanes=4 inside MAX_DRAWABLE_HALF_ROAD_FT.
        "laneWidth": 10,
    }
    placements, params = _pipeline(body)
    expected = f"Shoulder Closure — {2 * lanes}-Lane Undivided"

    assert scenario_display_name(params) == expected
    # Plan-sheet title block MHT TYPE row renders this wrapper.
    assert _scenario_label(params) == expected.upper()
    # PARAMETERS box short form drops the qualifier entirely.
    assert scenario_display_name_short(params) == "Shoulder Closure"
    # Crew narrative header.
    assert build_narrative_context(placements, params)["closure_type_display"] == expected

    # XLSX device list Summary sheet.
    xlsx_path = tmp_path / "devices.xlsx"
    export_device_list(placements, params, str(xlsx_path))
    wb = load_workbook(str(xlsx_path), read_only=True)
    summary = {str(r[0]): r[1] for r in wb["Summary"].iter_rows(values_only=True)}
    wb.close()
    assert summary["Closure type"] == expected

    # Quote Summary header line.
    quote_path = tmp_path / "quote.xlsx"
    generate_quote(placements, params, output_path=str(quote_path))
    qwb = load_workbook(str(quote_path), read_only=True)
    header_row = next(
        qwb["Quote Summary"].iter_rows(min_row=6, max_row=6, max_col=1, values_only=True)
    )
    qwb.close()
    assert f"Closure: {expected}" in str(header_row[0])


def test_flagger_label_unchanged_by_lane_count_derivation() -> None:
    """The bridge forces flagger to num_lanes=1 (schemas.scenario_to_call),
    so the derived label is byte-identical to the former literal."""
    scenario = FlaggerLaneClosureScenario.model_validate(
        {
            "kind": "flagger_lane_closure",
            "roadType": "rural_undivided",
            "speed": 45,
            "laneWidth": 12,
            "workType": "utility_cut",
            "duration": "short",
            "workLen": 500,
            "night": False,
            "pilotCar": False,
            "afad": False,
            "pedestrianAccess": False,
        }
    )
    params, _generator, _kwargs = scenario_to_call(scenario)
    assert params.num_lanes == 1
    assert scenario_display_name(params) == "Flagger Alternating Traffic — 2-Lane Undivided"


def test_flagger_label_states_drawn_geometry_not_input_count() -> None:
    """The flagger label is a literal, not 2 * num_lanes (#117 enablement
    item): generate_flagger_alternating_2lane draws a 2-lane road
    unconditionally, so a direct caller passing num_lanes=2 must not get
    a "4-Lane Undivided" claim about a plan nobody drew.  Unreachable
    through the wire (the bridge forces num_lanes=1) — this pins the
    direct-construction path the old formula lied on."""
    params = ScenarioParams(
        speed_mph=45,
        num_lanes=2,
        lane_width_ft=12.0,
        closure_type="lane",
        road_type="rural",
        work_zone_length_ft=500.0,
        is_divided=False,
        jurisdiction="CDOT",
    )
    assert scenario_display_name(params) == "Flagger Alternating Traffic — 2-Lane Undivided"


def test_narrative_rural_road_type_makes_no_lane_claim() -> None:
    """ "Rural two-lane" asserted a lane count Table 6B-1 does not carry —
    the narrative road-type row must stay count-free (Refs #118)."""
    placements, params = _pipeline(
        {**CASE_11_GENERAL_BODY, "roadType": "rural_undivided", "divided": False, "lanes": 3}
    )
    assert params.road_type == "rural"
    assert build_narrative_context(placements, params)["road_type_human"] == "Rural"
