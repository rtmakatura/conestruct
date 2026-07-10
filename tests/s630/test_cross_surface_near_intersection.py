"""Cross-surface agreement for the (gated) near_intersection kind.

Option C increment 2 (Refs #117).  The quote, XLSX device list, and UI
device breakdown are approach-blind reductions over the placement list —
so the cross-street advance sets must ride into all three surfaces
automatically, with real catalog descriptions (no bare-code fallbacks,
no zero-priced unknowns) and byte-identical rows across surfaces
(the #101 invariant, extended to the new kind).

The kind is gated: placements come from the internal call path
(``scenario_to_call`` → generator), exactly like the other gated-kind
tests.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl import load_workbook

from src.api.render_api import _build_device_breakdown
from src.api.schemas import NearIntersectionScenario, scenario_to_call
from src.export.device_list import export_device_list
from src.export.quote_generator import generate_quote
from src.rules.validators import DevicePlacement, ScenarioParams

_TEMPLATE_TOKEN = re.compile(r"\bX{2,3}\b")

# The Case 18 anchor scenario with BOTH cross-street legs, so every
# cross-surface count exercises the per-approach multiplication.
BODY: dict[str, Any] = {
    "kind": "near_intersection",
    "meta": {
        "project": "ni-cross-surface",
        "address": "",
        "lat": 0.0,
        "lng": 0.0,
        "siteConditions": {},
    },
    "roadType": "urban_arterial",
    "speed": 45,
    "lanes": 2,
    "laneWidth": 12.0,
    "divided": False,
    "workType": "utility_cut",
    "duration": "short",
    "workLen": 300.0,
    "night": False,
    "approaches": [
        {
            "id": "cross_n",
            "speed": 30,
            "roadType": "urban_arterial",
            "lanesPerDirection": 1,
            "laneWidth": 12.0,
            "signalized": False,
            "bearingDeg": None,
            "alongStationFt": -60.0,
        },
        {
            "id": "cross_s",
            "speed": 30,
            "roadType": "urban_arterial",
            "lanesPerDirection": 1,
            "laneWidth": 12.0,
            "signalized": True,  # same set regardless (flag-and-cite is inc. 3)
            "bearingDeg": None,
            "alongStationFt": -60.0,
        },
    ],
}


def _pipeline() -> tuple[list[DevicePlacement], ScenarioParams]:
    scenario = NearIntersectionScenario.model_validate(BODY)
    params, generator, kwargs = scenario_to_call(scenario)
    return generator(params, **kwargs), params


def _xlsx_device_rows(placements, params, tmp_path) -> list[tuple]:
    path = tmp_path / "devices.xlsx"
    export_device_list(placements, params, str(path))
    wb = load_workbook(str(path), read_only=True)
    rows = list(wb["Device List"].iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _quote_equipment_rows(placements, params, tmp_path) -> list[tuple]:
    path = tmp_path / "quote.xlsx"
    generate_quote(placements, params, output_path=str(path))
    wb = load_workbook(str(path), read_only=True)
    rows = [
        r
        for r in wb["Equipment Detail"].iter_rows(min_row=2, values_only=True)
        if isinstance(r[0], int)
    ]
    wb.close()
    return rows


def test_quote_rows_match_xlsx_order_and_descriptions(tmp_path) -> None:
    """Byte-identical descriptions and quantities, same order, signs leading."""
    placements, params = _pipeline()
    quote_rows = _quote_equipment_rows(placements, params, tmp_path)
    xlsx_rows = _xlsx_device_rows(placements, params, tmp_path)
    assert quote_rows, "quote has no equipment rows"
    assert [str(r[3]) for r in quote_rows] == [str(r[2]) for r in xlsx_rows]
    assert [r[4] for r in quote_rows] == [r[5] for r in xlsx_rows]
    assert quote_rows[0][1] == "SIGN_GENERIC"


def test_breakdown_quantities_match_xlsx(tmp_path) -> None:
    placements, params = _pipeline()
    xlsx_rows = _xlsx_device_rows(placements, params, tmp_path)
    breakdown = _build_device_breakdown(placements, params)
    assert sorted(r["qty"] for r in breakdown) == sorted(r[5] for r in xlsx_rows)
    assert sum(r["qty"] for r in breakdown) == len(placements)


def test_cross_street_devices_are_counted_and_multiplied(tmp_path) -> None:
    """The per-approach sets ride into the paperwork: with two legs, the
    W20-1 / R2-10 / R2-11 quantities are mainline + 2 (Option C's whole
    point — counted and priced, not hand-added)."""
    placements, params = _pipeline()

    def qty(label: str) -> int:
        return sum(1 for p in placements if p.label == label)

    # 1 mainline + 1 per leg.
    assert qty("W20-1") == 3
    assert qty("R2-10") == 3
    assert qty("R2-11") == 3

    xlsx_rows = _xlsx_device_rows(placements, params, tmp_path)
    by_code_prefix = {}
    for r in xlsx_rows:
        desc = str(r[2])
        code = desc.split(" ", 1)[0]
        by_code_prefix[code] = by_code_prefix.get(code, 0) + int(r[5])
    assert by_code_prefix.get("W20-1") == 3
    assert by_code_prefix.get("R2-10") == 3
    assert by_code_prefix.get("R2-11") == 3


def test_every_sign_has_a_real_catalog_description(tmp_path) -> None:
    """Rule 10: an unknown sign code silently falls back to its bare code
    in SIGN_DESCRIPTIONS — every code this generator emits must resolve
    to a real description on every surface, and no description may ship
    a template token."""
    placements, params = _pipeline()
    for row in _quote_equipment_rows(placements, params, tmp_path):
        description = str(row[3])
        assert not _TEMPLATE_TOKEN.search(description)
        if row[1] == "SIGN_GENERIC" and row[2]:
            assert description != str(row[2]), (
                f"sign fell back to its bare code (missing SIGN_DESCRIPTIONS row): {description!r}"
            )
    for row in _build_device_breakdown(placements, params):
        assert not _TEMPLATE_TOKEN.search(str(row["device"]))


def test_no_zero_priced_equipment_rows(tmp_path) -> None:
    """No cross-street (or mainline) device may price at zero — a missing
    rate must fail loudly upstream, never ship a $0 row."""
    placements, params = _pipeline()
    for row in _quote_equipment_rows(placements, params, tmp_path):
        daily_rate = row[6]
        assert isinstance(daily_rate, (int, float)) and daily_rate > 0, (
            f"zero/absent daily rate on quote row: {row!r}"
        )
