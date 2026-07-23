"""The priced quote joins the shared delta-aware aggregation (issue #154).

Before #154 the quote workbook priced raw placements through its own
``_aggregate`` and never saw jurisdiction count-deltas, so a fired delta moved
the XLSX bid document and the on-sheet summary but NOT the contractor estimate.
These tests drive the real ``/render/quote`` and ``/render/quote-breakdown``
endpoints so a fired delta is proven to reach the rendered Equipment Detail
sheet and the quote totals:

  * a mapped delta (Denver's arrow board) is priced from EQUIPMENT_DAILY_RATES
    and moves the grand total by exactly its marked-up extended amount;
  * an unmapped delta (Thornton's advance_warning_signs) renders an honest "—"
    line that is EXCLUDED from the total, never a fabricated $0 rate
    (Option C, 2026-07-22 ruling);
  * a scenario with no jurisdiction is untouched (zero-churn plumbing).
"""

from __future__ import annotations

import copy
import io
import os

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.api.render_api import app
from tests.test_plan_sheet_device_summary import FLAGGER_BODY

# Em-dash used for the honest unmapped rate/extended cells and the XLSX
# pay-item convention (device_list uses the same glyph).
_DASH = "—"


@pytest.fixture(scope="module", autouse=True)
def _set_render_secret() -> None:
    os.environ["RENDER_API_SECRET"] = "test-secret-do-not-deploy"


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app)


def _auth() -> dict[str, str]:
    return {"Authorization": "Bearer test-secret-do-not-deploy"}


def _equipment_rows(
    client: TestClient, scenario: dict, settings: dict | None = None
) -> list[tuple]:
    """Equipment Detail data rows from a real /render/quote render."""
    payload: dict = {"scenario": scenario}
    if settings is not None:
        payload["settings"] = settings
    resp = client.post("/render/quote", json=payload, headers=_auth())
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    rows = [
        r
        for r in wb["Equipment Detail"].iter_rows(min_row=2, values_only=True)
        if isinstance(r[0], int)
    ]
    wb.close()
    return rows


def _breakdown(client: TestClient, scenario: dict, settings: dict | None = None) -> dict:
    payload: dict = {"scenario": scenario}
    if settings is not None:
        payload["settings"] = settings
    resp = client.post("/render/quote-breakdown", json=payload, headers=_auth())
    assert resp.status_code == 200, resp.text
    return resp.json()


# Equipment Detail columns:
# 0 Item# 1 Device Type 2 Label 3 Description 4 Qty 5 Unit 6 Daily Rate 7 Days
# 8 Extended 9 Notes


def test_denver_flagger_arrow_board_prices_on_quote_workbook(client: TestClient) -> None:
    """A fired Denver arrow-board delta is a priced Equipment Detail line, at
    the catalog daily rate ($85), with a jurisdiction-required note."""
    scenario = copy.deepcopy(FLAGGER_BODY)
    scenario["jurisdiction_key"] = "denver"
    scenario["street_class"] = "arterial"

    rows = _equipment_rows(client, scenario)
    arrow = [r for r in rows if r[1] == "Arrow Board"]
    assert len(arrow) == 1, [r[1] for r in rows]
    a = arrow[0]
    assert a[4] == 1  # qty
    assert a[5] == "EACH"
    assert a[6] == 85.0  # daily rate, sourced from EQUIPMENT_DAILY_RATES
    assert a[8] == 85.0  # extended = 1 x 85 x 1 day (default duration)
    assert "Jurisdiction-required" in str(a[9])


def test_denver_flagger_total_moves_by_marked_up_arrow_board(client: TestClient) -> None:
    """The grand total moves by EXACTLY the marked-up arrow-board line and
    nothing else — equipment_total up by its rate, total up by the overhead/
    profit-expanded amount."""
    base_scenario = copy.deepcopy(FLAGGER_BODY)
    denver = copy.deepcopy(FLAGGER_BODY)
    denver["jurisdiction_key"] = "denver"
    denver["street_class"] = "arterial"

    b = _breakdown(client, base_scenario)
    d = _breakdown(client, denver)

    equip_delta = d["equipment_total"] - b["equipment_total"]
    assert equip_delta == pytest.approx(85.0)  # 1 x $85 x 1 day

    # Labor and delivery are untouched by a device delta, so the whole total
    # move is the marked-up equipment delta.
    expected_total_delta = equip_delta * (1 + d["overhead_pct"]) * (1 + d["profit_pct"])
    assert d["total"] - b["total"] == pytest.approx(expected_total_delta)


def test_thornton_night_unmapped_advance_warning_excluded_from_total(client: TestClient) -> None:
    """Thornton night lane fires a drum top-up (real rate, quantity unchanged)
    and advance_warning_signs (unmapped): the unmapped line renders "—" rate
    and extended, states it is not in the total, and the equipment total is
    unchanged from the same scenario without the jurisdiction."""
    scenario = copy.deepcopy(FLAGGER_BODY)
    scenario["jurisdiction_key"] = "thornton"
    scenario["street_class"] = "arterial"
    scenario["night"] = True

    rows = _equipment_rows(client, scenario)

    # Drum: topped up to the layout's 6 (max(6, 1)), priced normally, note added.
    drum = [r for r in rows if str(r[1]).upper() == "DRUM"]
    assert len(drum) == 1
    assert drum[0][4] == 6  # qty unchanged
    assert drum[0][6] == 3.0  # real daily rate
    assert "Jurisdiction-required" in str(drum[0][9])

    # advance_warning_signs: unmapped -> honest "—" line, excluded from total.
    aws = [r for r in rows if str(r[1]).startswith("Advance Warning Signs")]
    assert len(aws) == 1
    assert aws[0][6] == _DASH  # daily rate cell, not $0.00
    assert aws[0][8] == _DASH  # extended cell, not $0.00
    assert "NOT included in the quote total" in str(aws[0][9])
    assert "price separately" in str(aws[0][9])

    # The grand-total math excludes the unmapped line: equipment_total equals
    # the same night scenario without the jurisdiction.
    base = copy.deepcopy(FLAGGER_BODY)
    base["night"] = True
    b = _breakdown(client, base)
    d = _breakdown(client, scenario)
    assert d["equipment_total"] == pytest.approx(b["equipment_total"])


def test_no_jurisdiction_quote_is_inert(client: TestClient) -> None:
    """Zero-churn: a scenario naming no jurisdiction carries no jurisdiction
    provenance on any equipment line — the delta path is a no-op (rule 5)."""
    b = _breakdown(client, copy.deepcopy(FLAGGER_BODY))
    for line in b["equipment_lines"]:
        assert line["jurisdiction_required"] is False, line
        assert line["jurisdiction_unmapped"] is False, line
