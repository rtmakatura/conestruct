"""Spec §4 cross-surface proof (issue #150/#151): the sheet's summary IS the
XLSX's rows IS the screen breakdown, fired jurisdiction count-deltas included,
and required-on-sheet jurisdictions cannot toggle the block off.

Issue #151 (Ryan ruling, 2026-07-22) moved jurisdiction count-deltas into the
shared aggregation, so a fired delta now reaches the XLSX bid document and the
on-sheet summary — not only the screen.  The parity tests below assert the
delta-adjusted quantities on all three surfaces; the pre-#151 caveat (parity
against no-delta XLSX behavior) is gone.
"""

from __future__ import annotations

import copy
import io

import pytest
from openpyxl import load_workbook

from tests.test_plan_sheet_device_summary import (
    FLAGGER_BODY,
    SHOULDER_BODY,
    _placements_for_body,
    pdf_text,
)
from tests.test_plan_sheet_device_summary import (
    _secret as _secret,  # noqa: PLC0414 — autouse env fixture
)
from tests.test_plan_sheet_device_summary import (
    client as client,  # noqa: PLC0414 — TestClient fixture
)

AUTH = {"Authorization": "Bearer test-secret"}


def _xlsx_device_rows(client, body):
    """(Device-List data rows, Summary 'Total device count') from /render/xlsx."""
    resp = client.post("/render/xlsx", json=body, headers=AUTH)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    rows = [r for r in wb["Device List"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    total = next(
        (
            r[1]
            for r in wb["Summary"].iter_rows(min_row=2, values_only=True)
            if r[0] == "Total device count"
        ),
        None,
    )
    return rows, total


def _breakdown_total(client, body):
    resp = client.post("/render/device-breakdown", json=body, headers=AUTH)
    assert resp.status_code == 200, resp.text
    return resp.json()["total_devices"]


@pytest.mark.parametrize("body", [SHOULDER_BODY, FLAGGER_BODY], ids=["shoulder", "flagger"])
def test_every_xlsx_row_quantity_appears_on_sheet(client, tmp_path, body) -> None:
    from src.rules.jurisdiction import aggregate_device_rows_with_deltas

    body = copy.deepcopy(body)
    placements, _params = _placements_for_body(body)
    rows = aggregate_device_rows_with_deltas(placements, [])
    text = pdf_text(client, body, tmp_path)
    for row in rows:
        assert str(row.quantity) in text, (row.device_type, row.label)
    assert str(sum(r.quantity for r in rows)) in text  # totals row


def test_greeley_arterial_flagger_delta_reaches_all_three_surfaces(client, tmp_path) -> None:
    """A fired Greeley Type-C arrow-board delta (arterial lane closure) appears
    on the XLSX bid line with its real CDOT pay item, on the on-sheet summary,
    and the three surfaces agree on the total (issue #151)."""
    body = copy.deepcopy(FLAGGER_BODY)
    body["jurisdiction_key"] = "greeley"
    body["street_class"] = "arterial"

    rows, xlsx_total = _xlsx_device_rows(client, body)
    arrow = [r for r in rows if r[1] == "Arrow Board (Type C)"]
    assert len(arrow) == 1, [r[1] for r in rows]
    # (Item#, DeviceType, Description, PayItem, Unit, Qty, Notes)
    assert arrow[0][3] == "630-80358"  # real CDOT C-type pay item, not fabricated
    assert arrow[0][4] == "EACH"
    assert arrow[0][5] == 1
    assert "Jurisdiction-required" in arrow[0][6]

    # On-sheet summary shows the same device and total.
    text = pdf_text(client, body, tmp_path)
    assert "Arrow Board (Type C)" in text

    # Cross-surface equality: breakdown == XLSX summary == sum of XLSX qtys.
    xlsx_qty_sum = sum(int(r[5]) for r in rows)
    assert _breakdown_total(client, body) == xlsx_total == xlsx_qty_sum
    assert str(xlsx_total) in text  # totals row on the sheet


def test_thornton_night_lane_multi_delta_maps_and_honest_unmapped(client, tmp_path) -> None:
    """Thornton night lane closure fires two count deltas: a drum top-up (real
    pay item, quantity unchanged) and advance_warning_signs — an unmapped
    device that renders as an honest, pay-item-less bid line rather than a
    fabricated one (rule 10 / issue #151)."""
    body = copy.deepcopy(FLAGGER_BODY)
    body["jurisdiction_key"] = "thornton"
    body["street_class"] = "arterial"
    body["night"] = True

    rows, xlsx_total = _xlsx_device_rows(client, body)

    # Drum topped up, NOT added: the flagger night-lane layout already carries
    # 6 drums and the delta requires >= 1, so max(6, 1) == 6 — the row stays a
    # single DRUM row at quantity 6 (proving no double-count on the bid
    # document) and only gains the jurisdiction-required note.
    drum = [r for r in rows if r[3] == "630-80360"]
    assert len(drum) == 1
    assert drum[0][5] == 6  # (Item#, DeviceType, Desc, PayItem, Unit, Qty, Notes)
    assert drum[0][4] == "EACH"
    assert "Jurisdiction-required" in (drum[0][6] or "")

    # advance_warning_signs: unmapped -> no fabricated pay item, honest note.
    aws = [r for r in rows if str(r[1]).startswith("Advance Warning Signs")]
    assert len(aws) == 1
    assert aws[0][3] == "—"  # no single CDOT pay item
    assert aws[0][5] == 1
    assert "no single pay item" in aws[0][6]
    assert "jurisdiction-required" in aws[0][1]

    # Cross-surface equality holds with the multi-delta config too.
    text = pdf_text(client, body, tmp_path)
    xlsx_qty_sum = sum(int(r[5]) for r in rows)
    assert _breakdown_total(client, body) == xlsx_total == xlsx_qty_sum
    assert str(xlsx_total) in text


def test_required_jurisdiction_ignores_toggle_off(client, tmp_path) -> None:
    body = copy.deepcopy(SHOULDER_BODY)
    body["jurisdiction_key"] = "loveland"
    body["meta"]["includeDeviceSummary"] = False
    assert "TRAFFIC CONTROL DEVICE SUMMARY" in pdf_text(client, body, tmp_path)


def test_optional_jurisdiction_honors_toggle_off(client, tmp_path) -> None:
    body = copy.deepcopy(SHOULDER_BODY)
    body["jurisdiction_key"] = "cdot"
    body["meta"]["includeDeviceSummary"] = False
    assert "TRAFFIC CONTROL DEVICE SUMMARY" not in pdf_text(client, body, tmp_path)


def test_wave2_gate_trio_all_render_the_block(client, tmp_path) -> None:
    """The BLOCKED.md hard gate, as an executable statement."""
    for key in ("castle_rock", "loveland", "thornton"):
        body = copy.deepcopy(SHOULDER_BODY)
        body["jurisdiction_key"] = key
        body["meta"]["includeDeviceSummary"] = False  # even opted out
        assert "TRAFFIC CONTROL DEVICE SUMMARY" in pdf_text(client, body, tmp_path), key
