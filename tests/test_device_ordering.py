"""Device-list ordering invariants (issue #88).

The aggregated device surfaces — the UI "Plan details" panel
(``_build_device_breakdown``), the XLSX Device List
(``_populate_device_list_sheet``), and the crew-narrative Required
Equipment list (``_device_summary``) — all derive their row order from
the single shared :func:`device_row_sort_key` helper.  These tests pin
that helper's grouped order (signs → channelizing → equipment, each
group sorted as specified), guard the category map's completeness, and
assert the three surfaces order an identical device set identically.

Reordering rows changes no count: the per-surface count assertions in
``tests/s630/test_cross_surface.py`` and the audit snapshots (sorted-key
dicts, order-independent) are the separate guard on that.
"""

from __future__ import annotations

import os
import tempfile

from openpyxl import load_workbook

from src.api.render_api import _NON_SIGN_DISPLAY, _build_device_breakdown
from src.export.device_list import export_device_list
from src.generation.layout import generate_shoulder_closure_divided
from src.narrative.crew_narrative import _DEVICE_HUMAN_NAMES, build_narrative_context
from src.rules.devices import (
    DeviceCategory,
    DeviceType,
    cone_display_name,
    device_category,
    device_row_sort_key,
)
from src.rules.validators import ScenarioParams

_PARAMS = ScenarioParams(
    speed_mph=65,
    num_lanes=2,
    closure_type="shoulder",
    road_type="freeway",
    work_zone_length_ft=800.0,
    is_divided=True,
    jurisdiction="CDOT",
)


# ---------------------------------------------------------------------------
# Category map completeness — a new DeviceType must be categorized
# ---------------------------------------------------------------------------


def test_every_device_type_has_a_category() -> None:
    """Adding a DeviceType without categorizing it should fail loudly here,
    not silently sort the device to an arbitrary position."""
    for dt in DeviceType:
        # Raises KeyError (caught as a test failure) if dt is uncategorized.
        assert isinstance(device_category(dt), DeviceCategory), dt


def test_every_non_sign_device_has_a_sort_name() -> None:
    """Every non-sign device must have a canonical sort name; only
    SIGN_GENERIC is exempt (signs order by schedule key)."""
    for dt in DeviceType:
        if dt is DeviceType.SIGN_GENERIC:
            continue
        # device_row_sort_key reads _DEVICE_SORT_NAME for non-signs; a
        # missing entry raises KeyError.
        key = device_row_sort_key(dt, None)
        assert key[0] != DeviceCategory.SIGN.value, dt
        assert isinstance(key[2], str) and key[2], dt


# ---------------------------------------------------------------------------
# Helper order — signs first, then channelizing, then equipment
# ---------------------------------------------------------------------------


def test_device_row_sort_key_grouped_order() -> None:
    """Pin the exact grouped order the helper produces for a representative
    mix: signs (by schedule key, unlabeled last), then channelizing devices
    and equipment, each alphabetical by canonical display name."""
    rows = [
        (DeviceType.PCMS, None),
        (DeviceType.SIGN_GENERIC, "W20-1"),
        (DeviceType.BARRICADE_TYPE_II, None),
        (DeviceType.CONE, None),
        (DeviceType.SIGN_GENERIC, None),  # unlabeled sign
        (DeviceType.ARROW_BOARD, None),
        (DeviceType.DRUM, None),
        (DeviceType.SIGN_GENERIC, "G20-2"),
        (DeviceType.TEMPORARY_BARRIER, None),
    ]
    ordered = sorted(rows, key=lambda r: device_row_sort_key(r[0], r[1]))
    assert ordered == [
        # signs: labeled by schedule key, then unlabeled
        (DeviceType.SIGN_GENERIC, "G20-2"),
        (DeviceType.SIGN_GENERIC, "W20-1"),
        (DeviceType.SIGN_GENERIC, None),
        # channelizing devices, alphabetical by display name
        (DeviceType.DRUM, None),  # "Channelizing Drum"
        (DeviceType.TEMPORARY_BARRIER, None),  # "Temporary Barrier"
        (DeviceType.CONE, None),  # "Traffic Cone"
        (DeviceType.BARRICADE_TYPE_II, None),  # "Type II Barricade"
        # equipment, alphabetical by display name
        (DeviceType.ARROW_BOARD, None),  # "Arrow Board"
        (DeviceType.PCMS, None),  # "Portable Changeable Message Sign"
    ]


# ---------------------------------------------------------------------------
# Cross-surface invariant — all three surfaces order the set identically
# ---------------------------------------------------------------------------


def _panel_non_sign_types() -> list[DeviceType]:
    """Non-sign DeviceTypes in panel order (rows with code '—')."""
    placements = generate_shoulder_closure_divided(_PARAMS)
    inverse = {disp: dt for dt, (disp, _fn) in _NON_SIGN_DISPLAY.items()}
    inverse[cone_display_name(_PARAMS.speed_mph)] = DeviceType.CONE
    out: list[DeviceType] = []
    for row in _build_device_breakdown(placements, _PARAMS):
        if row["code"] == "—":
            out.append(inverse[row["device"]])
    return out


def _crew_non_sign_types() -> list[DeviceType]:
    """Non-sign DeviceTypes in crew-equipment order (top-level rows with no
    nested children — the signs are nested under one parent)."""
    placements = generate_shoulder_closure_divided(_PARAMS)
    context = build_narrative_context(placements, _PARAMS)
    inverse = {disp: dt for dt, disp in _DEVICE_HUMAN_NAMES.items()}
    inverse[cone_display_name(_PARAMS.speed_mph)] = DeviceType.CONE
    out: list[DeviceType] = []
    for row in context["device_summary"]:
        if not row["children"]:
            out.append(inverse[row["label"]])
    return out


def _xlsx_rows() -> list[tuple]:
    placements = generate_shoulder_closure_divided(_PARAMS)
    tmp = tempfile.mkdtemp()
    path = os.path.join(tmp, "devices.xlsx")
    export_device_list(placements, _PARAMS, path)
    wb = load_workbook(path, read_only=True)
    rows = list(wb["Device List"].iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _xlsx_non_sign_types() -> list[DeviceType]:
    """Non-sign DeviceTypes in XLSX order.  The Device Type column carries
    the enum value for non-signs, so the mapping is exact."""
    out: list[DeviceType] = []
    for row in _xlsx_rows():
        type_label = str(row[1])
        if not type_label.startswith("SIGN_GENERIC"):
            out.append(DeviceType(type_label))
    return out


def test_three_surfaces_order_non_signs_identically() -> None:
    """The panel, the XLSX device list, and the crew equipment list emit the
    same non-sign device set in the same order — the order the shared helper
    dictates."""
    panel = _panel_non_sign_types()
    crew = _crew_non_sign_types()
    xlsx = _xlsx_non_sign_types()

    expected = sorted(set(panel), key=lambda dt: device_row_sort_key(dt, None))

    assert panel == expected, panel
    assert crew == expected, crew
    assert xlsx == expected, xlsx


def test_signs_lead_every_surface() -> None:
    """Signs are the first group on all three surfaces (issue #88)."""
    placements = generate_shoulder_closure_divided(_PARAMS)

    # Panel: first row is a sign (has a real MUTCD code, not the '—' filler).
    panel = _build_device_breakdown(placements, _PARAMS)
    assert panel[0]["code"] != "—", panel[0]

    # XLSX: first data row is a sign.
    assert str(_xlsx_rows()[0][1]).startswith("SIGN_GENERIC")

    # Crew: first top-level bullet is the sign parent (it carries children).
    summary = build_narrative_context(placements, _PARAMS)["device_summary"]
    assert summary[0]["children"], summary[0]
