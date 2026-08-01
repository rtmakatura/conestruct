"""Build a CDOT-format pricing quote workbook from a placement list.

Produces a 5-sheet contractor estimate (summary, equipment, labor, delivery,
terms) with line-item rates, overhead, and profit applied. Pricing is
calibrated to Colorado Front Range traffic-control vendors as of 2026.
Final pricing is subject to site inspection and signed contract terms.

Authoritative sources:
  - CDOT Standard Specifications Section 630 (pay item names, units)
  - Colorado labor law (overtime after 8 hours per shift)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.rules.device_aggregation import AggregatedDeviceRow
from src.rules.devices import DEVICE_CATALOG, DeviceType, cone_display_name
from src.rules.jurisdiction import aggregate_device_rows_with_deltas
from src.rules.sign_codes import display_code, substitute_sign_description
from src.rules.validators import DevicePlacement, ScenarioParams, scenario_display_name

# ---------------------------------------------------------------------------
# Pricing constants
# ---------------------------------------------------------------------------

EQUIPMENT_DAILY_RATES: dict[DeviceType, float] = {
    DeviceType.CONE: 1.50,
    DeviceType.DRUM: 3.00,
    DeviceType.TUBULAR_MARKER: 1.00,
    DeviceType.BARRICADE_TYPE_II: 5.00,
    DeviceType.BARRICADE_TYPE_III: 8.00,
    DeviceType.LONGITUDINAL_CHANNELIZER: 2.00,
    DeviceType.ARROW_BOARD: 85.00,
    DeviceType.PCMS: 175.00,
    DeviceType.TRUCK_MOUNTED_ATTENUATOR: 350.00,
    DeviceType.TEMPORARY_BARRIER: 15.00,
    DeviceType.FLAGGER_STATION: 0.00,
    DeviceType.TEMPORARY_SIGNAL: 200.00,
    DeviceType.SIGN_GENERIC: 4.00,
    DeviceType.DETOUR_MARKER: 4.00,
    DeviceType.CHANNELIZER_OPTIONAL: 2.00,
    # TODO: replace placeholder rates with CBC actual pricing once a
    # nighttime contractor invoice is available.
    DeviceType.WARNING_LIGHT_TYPE_C: 5.00,
    DeviceType.PORTABLE_LIGHT_PLANT: 125.00,
}

_FLAGGER_NOTE = "See Labor Detail"
_CHANNELIZER_OPTIONAL_NOTE = "Optional — quantities may be reduced per engineer discretion"

_LEFT_ALIGN: Alignment = Alignment(horizontal="left", vertical="center")


def _apply_left_align(sheet) -> None:
    """Left-align every populated cell on ``sheet``.

    Excel right-aligns numbers and left-aligns text by default; this
    forces a consistent visual alignment across columns.  Bold/fill
    on header cells is preserved (alignment is independent of font).
    """
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.alignment = _LEFT_ALIGN


# Standard traffic-control shift length per Colorado field practice — 8 hr
# work + 2 hr mobilization/demob = 10 hr billed.
_FLAGGER_HOURS_PER_DAY: float = 10.0

_TERMS: tuple[str, ...] = (
    "1. Prices valid for 30 days from date of estimate.",
    "2. Minimum 4-hour call-out for all labor.",
    "3. Night work (sunset to sunrise) billed at 1.5x standard rate.",
    "4. Equipment rental begins at time of delivery to site.",
    "5. Setup and takedown charged per occurrence.",
    "6. Damaged or missing equipment billed at replacement cost.",
    "7. Traffic control plan (MHT) included with this estimate.",
    "8. All work performed by ATSSA/CCA-certified personnel.",
    "9. This estimate is for planning purposes only. Final pricing subject "
    "to site inspection and contract terms.",
)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(fill_type="solid", start_color="FFF5F0EB", end_color="FFF5F0EB")
_HEADER_FONT = Font(bold=True, color="FF1B2838")
_SUBTOTAL_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True, size=13)
_TITLE_FONT = Font(bold=True, size=16)
_SUBTITLE_FONT = Font(bold=True, size=14)
_FOOTNOTE_FONT = Font(italic=True, size=9, color="FF888888")
_TOP_BORDER = Border(top=Side(border_style="thin", color="FF1B2838"))
_CURRENCY_FMT = "$#,##0.00"


# ---------------------------------------------------------------------------
# Breakdown data structures (returned alongside the file path)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class EquipmentLine:
    item_number: int
    device_type: str
    label: str
    description: str
    qty: int
    unit: str
    daily_rate: float
    days: int
    extended: float
    note: str
    # Jurisdiction count-delta provenance (issue #151/#154).  A row is
    # jurisdiction_required when a fired add_device delta topped it up or
    # added it.  jurisdiction_unmapped marks a delta device with no single
    # daily rate (e.g. advance_warning_signs): its daily_rate/extended are
    # 0.0 and excluded from the equipment total, and the sheet renders the
    # rate/extended cells as "—" (never $0.00) so the priced document can
    # never read the device as free (rule 10 / 2026-07-22 ruling).
    jurisdiction_required: bool = False
    jurisdiction_unmapped: bool = False


@dataclass(frozen=True)
class LaborLine:
    role: str
    personnel: int
    hours_per_day: float
    days: int
    rate: float
    extended: float


@dataclass(frozen=True)
class DeliveryLine:
    item: str
    trips: int
    distance_miles: float
    rate_per_mile: float
    min_trip_charge: float
    extended: float


@dataclass(frozen=True)
class QuoteBreakdown:
    equipment_lines: list[EquipmentLine] = field(default_factory=list)
    labor_lines: list[LaborLine] = field(default_factory=list)
    delivery_lines: list[DeliveryLine] = field(default_factory=list)
    is_night: bool = False
    night_multiplier: float = 1.0
    overhead_pct: float = 0.0
    profit_pct: float = 0.0
    equipment_total: float = 0.0
    labor_total: float = 0.0
    delivery_total: float = 0.0
    subtotal: float = 0.0
    overhead: float = 0.0
    profit: float = 0.0
    total: float = 0.0


# ---------------------------------------------------------------------------
# Jurisdiction count-delta notes (issue #151/#154)
# ---------------------------------------------------------------------------

# The quote joins the shared aggregation path (aggregate_device_rows_with_deltas)
# so a fired jurisdiction count delta is priced on the same rows the XLSX bid
# document and on-sheet summary render.  Mapped delta devices price from
# EQUIPMENT_DAILY_RATES via their catalog DeviceType; an unmapped device id
# (e.g. advance_warning_signs — a set of W-series signs with no single daily
# rate) renders an honest "—" line that is excluded from the total, never a
# fabricated rate (rule 10 / 2026-07-22 ruling).
_JURISDICTION_MAPPED_NOTE: str = "Jurisdiction-required — {doc}."
# Follows the shipped device_list._JURISDICTION_UNMAPPED_NOTE convention and
# reuses its "billed by SF" citation — no new citation.  The exclusion clause
# is binding (2026-07-22 ruling): a reader must be able to tell from the sheet
# that the grand total does NOT include this line.
_JURISDICTION_UNMAPPED_NOTE: str = (
    "Jurisdiction-required — {doc}. No single daily rate — NOT included in the "
    "quote total; price separately. CDOT Spec 630 bills these as individual "
    "W-series signs by SF — itemize per sign type."
)


def _cents(x: float) -> float:
    """#135 ruling (Refs #185): money rounds to cents at computation;
    every downstream sum consumes rounded values.  Half-cent cases round
    half away from zero — CHOSEN (Rule 12): standard invoice convention,
    not sourced from a spec — via Decimal(str(x)) so the decision is made
    on the decimal value, not on its binary float representation."""
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _jurisdiction_doc(source: dict[str, Any] | None) -> str:
    """Human doc name from a delta ``source`` block, for the Notes column."""
    return (source or {}).get("doc", "jurisdiction requirement")


def _description_for(
    device_type: DeviceType,
    label: str | None,
    representative: DevicePlacement | None,
    params: ScenarioParams,
) -> str:
    if device_type == DeviceType.SIGN_GENERIC:
        if label is None:
            return "Generic construction sign (unlabeled)"
        station_ft = representative.station_ft if representative is not None else 0.0
        code, human = substitute_sign_description(label, station_ft, params)
        return f"{code} {human}".strip() if human != code else code
    if device_type == DeviceType.CONE:
        return cone_display_name(params.speed_mph)
    return DEVICE_CATALOG[device_type].description


def _unit_for(device_type: DeviceType) -> str:
    if device_type == DeviceType.SIGN_GENERIC:
        return "EACH"
    return DEVICE_CATALOG[device_type].unit


def _note_for(device_type: DeviceType) -> str:
    if device_type == DeviceType.FLAGGER_STATION:
        return _FLAGGER_NOTE
    if device_type == DeviceType.CHANNELIZER_OPTIONAL:
        return _CHANNELIZER_OPTIONAL_NOTE
    return ""


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _style_header_row(sheet, row_index: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row_index, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _equipment_line_for(
    item_number: int,
    row: AggregatedDeviceRow,
    params: ScenarioParams,
    project_duration_days: int,
) -> EquipmentLine:
    """Price one aggregated device row into an equipment line.

    A jurisdiction delta-only add (no backing placement, ``display_override``
    set) prices from its mapped ``DeviceType`` — a real ``EQUIPMENT_DAILY_RATES``
    entry — or, for an unmapped device id, renders an honest "—" line excluded
    from the total (issue #151/#154).  A topped-up real row prices normally and
    gains a jurisdiction-required note.
    """
    device_type = row.device_type
    qty = row.quantity

    # Delta-only add (no backing placement).
    if row.display_override is not None:
        doc = _jurisdiction_doc(row.jurisdiction_source)
        if device_type is not None:
            rate = EQUIPMENT_DAILY_RATES.get(device_type, 0.0)
            spec = DEVICE_CATALOG[device_type]
            return EquipmentLine(
                item_number=item_number,
                device_type=row.display_override,
                label="",
                description=spec.description,
                qty=qty,
                unit=spec.unit,
                daily_rate=rate,
                days=project_duration_days,
                extended=_cents(qty * rate * project_duration_days),
                note=_JURISDICTION_MAPPED_NOTE.format(doc=doc),
                jurisdiction_required=True,
            )
        # Unmapped device id: no single daily rate.  Honest "—" line,
        # excluded from the total (daily_rate/extended stay 0.0 for the math;
        # the sheet renders "—").
        return EquipmentLine(
            item_number=item_number,
            device_type=f"{row.display_override} (jurisdiction-required)",
            label="",
            description=row.display_override,
            qty=qty,
            unit="EACH",
            daily_rate=0.0,
            days=project_duration_days,
            extended=0.0,
            note=_JURISDICTION_UNMAPPED_NOTE.format(doc=doc),
            jurisdiction_required=True,
            jurisdiction_unmapped=True,
        )

    # Normal row (possibly a delta-topped-up real row).
    label = row.label
    representative = row.representative
    rate = EQUIPMENT_DAILY_RATES.get(device_type, 0.0)
    note = _note_for(device_type)
    if row.jurisdiction_required:
        jur_note = _JURISDICTION_MAPPED_NOTE.format(doc=_jurisdiction_doc(row.jurisdiction_source))
        note = f"{note} {jur_note}".strip() if note else jur_note
    return EquipmentLine(
        item_number=item_number,
        device_type=device_type.value,
        # display_code strips the internal R2-1@face split marker
        # so both faces show "R2-1" (descriptions distinguish them).
        label=display_code(label) if label else "",
        description=_description_for(device_type, label, representative, params),
        qty=qty,
        unit=_unit_for(device_type),
        daily_rate=rate,
        days=project_duration_days,
        extended=_cents(qty * rate * project_duration_days),
        note=note,
        jurisdiction_required=row.jurisdiction_required,
    )


def _compute_equipment_lines(
    rows: list[AggregatedDeviceRow],
    params: ScenarioParams,
    project_duration_days: int,
) -> list[EquipmentLine]:
    return [
        _equipment_line_for(item_number, row, params, project_duration_days)
        for item_number, row in enumerate(rows, start=1)
    ]


def _populate_equipment_sheet(sheet, lines: list[EquipmentLine]) -> None:
    sheet.title = "Equipment Detail"
    headers = (
        "Item #",
        "Device Type",
        "Label",
        "Description",
        "Qty",
        "Unit",
        "Daily Rate",
        "Days",
        "Extended",
        "Notes",
    )
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    widths = {"A": 8, "B": 22, "C": 12, "D": 38, "E": 7, "F": 8, "G": 12, "H": 7, "I": 14, "J": 48}
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    for line in lines:
        # An unmapped jurisdiction row has no single daily rate: render "—"
        # (never $0.00) in the rate/extended cells so the priced document can
        # never read the device as free, and skip the currency format so the
        # dash shows literally (issue #151/#154).
        rate_cell = "—" if line.jurisdiction_unmapped else line.daily_rate
        extended_cell = "—" if line.jurisdiction_unmapped else line.extended
        sheet.append(
            (
                line.item_number,
                line.device_type,
                line.label,
                line.description,
                line.qty,
                line.unit,
                rate_cell,
                line.days,
                extended_cell,
                line.note,
            )
        )
        row_idx = line.item_number + 1
        if not line.jurisdiction_unmapped:
            sheet.cell(row=row_idx, column=7).number_format = _CURRENCY_FMT
            sheet.cell(row=row_idx, column=9).number_format = _CURRENCY_FMT

    equipment_total = sum(line.extended for line in lines)
    subtotal_row = len(lines) + 2
    sheet.cell(row=subtotal_row, column=8, value="SUBTOTAL").font = _SUBTOTAL_FONT
    cell = sheet.cell(row=subtotal_row, column=9, value=equipment_total)
    cell.font = _SUBTOTAL_FONT
    cell.number_format = _CURRENCY_FMT
    cell.border = _TOP_BORDER

    sheet.freeze_panes = "A2"


def _compute_labor_lines(
    *,
    is_night: bool,
    project_duration_days: int,
    num_flaggers: int,
    tcs_hours_per_day: float,
    num_setup_crew: int,
    setup_hours: float,
    takedown_hours: float,
    flagger_hourly_rate: float,
    tcs_hourly_rate: float,
    crew_hourly_rate: float,
    night_multiplier: float,
) -> list[LaborLine]:
    multiplier = night_multiplier if is_night else 1.0
    return [
        LaborLine(
            role="Flaggers",
            personnel=num_flaggers,
            hours_per_day=_FLAGGER_HOURS_PER_DAY,
            days=project_duration_days,
            rate=flagger_hourly_rate,
            extended=_cents(
                num_flaggers
                * _FLAGGER_HOURS_PER_DAY
                * project_duration_days
                * flagger_hourly_rate
                * multiplier
            ),
        ),
        LaborLine(
            role="TCS Supervisor",
            personnel=1,
            hours_per_day=tcs_hours_per_day,
            days=project_duration_days,
            rate=tcs_hourly_rate,
            extended=_cents(
                1 * tcs_hours_per_day * project_duration_days * tcs_hourly_rate * multiplier
            ),
        ),
        LaborLine(
            role="Setup Crew",
            personnel=num_setup_crew,
            hours_per_day=setup_hours,
            days=1,
            rate=crew_hourly_rate,
            extended=_cents(num_setup_crew * setup_hours * 1 * crew_hourly_rate * multiplier),
        ),
        LaborLine(
            role="Takedown Crew",
            personnel=num_setup_crew,
            hours_per_day=takedown_hours,
            days=1,
            rate=crew_hourly_rate,
            extended=_cents(num_setup_crew * takedown_hours * 1 * crew_hourly_rate * multiplier),
        ),
    ]


def _populate_labor_sheet(
    sheet,
    lines: list[LaborLine],
    *,
    is_night: bool,
    night_multiplier: float,
) -> None:
    sheet.title = "Labor Detail"
    headers = ("Role", "Personnel", "Hours/Day", "Days", "Rate ($/hr)", "Extended")
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    widths = {"A": 22, "B": 12, "C": 12, "D": 8, "E": 14, "F": 14}
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    for line in lines:
        sheet.append(
            (line.role, line.personnel, line.hours_per_day, line.days, line.rate, line.extended)
        )
        last = sheet.max_row
        sheet.cell(row=last, column=5).number_format = _CURRENCY_FMT
        sheet.cell(row=last, column=6).number_format = _CURRENCY_FMT

    if is_night:
        sheet.append(
            (
                "Night Differential",
                "",
                "",
                "",
                f"{night_multiplier:.1f}x",
                "(applied to all labor lines above)",
            )
        )

    labor_total = sum(line.extended for line in lines)
    subtotal_row_idx = sheet.max_row + 1
    sheet.cell(row=subtotal_row_idx, column=5, value="SUBTOTAL").font = _SUBTOTAL_FONT
    cell = sheet.cell(row=subtotal_row_idx, column=6, value=labor_total)
    cell.font = _SUBTOTAL_FONT
    cell.number_format = _CURRENCY_FMT
    cell.border = _TOP_BORDER

    note_row = subtotal_row_idx + 2
    sheet.cell(
        row=note_row,
        column=1,
        value=(
            "Standard shift: 10 hours including mobilization. "
            "Overtime billed at 1.5x after 8 hours per Colorado labor law."
        ),
    ).font = _FOOTNOTE_FONT
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)

    sheet.freeze_panes = "A2"


def _compute_delivery_lines(
    *,
    delivery_trips: int,
    delivery_distance_miles: float,
    mileage_rate: float,
    min_trip_charge: float,
) -> list[DeliveryLine]:
    per_trip = _cents(max(delivery_distance_miles * mileage_rate, min_trip_charge))
    extended = _cents(delivery_trips * per_trip)
    return [
        DeliveryLine(
            item="Equipment Delivery",
            trips=delivery_trips,
            distance_miles=delivery_distance_miles,
            rate_per_mile=mileage_rate,
            min_trip_charge=min_trip_charge,
            extended=extended,
        )
    ]


def _populate_delivery_sheet(sheet, lines: list[DeliveryLine]) -> None:
    sheet.title = "Delivery & Logistics"
    headers = ("Item", "Trips", "Distance (mi)", "Rate ($/mi)", "Min Charge", "Extended")
    sheet.append(headers)
    _style_header_row(sheet, 1, len(headers))

    widths = {"A": 28, "B": 8, "C": 14, "D": 14, "E": 14, "F": 14}
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    for line in lines:
        sheet.append(
            (
                line.item,
                line.trips,
                line.distance_miles,
                line.rate_per_mile,
                line.min_trip_charge,
                line.extended,
            )
        )
        last = sheet.max_row
        for col in (4, 5, 6):
            sheet.cell(row=last, column=col).number_format = _CURRENCY_FMT

    delivery_total = sum(line.extended for line in lines)
    subtotal_row = sheet.max_row + 1
    sheet.cell(row=subtotal_row, column=5, value="SUBTOTAL").font = _SUBTOTAL_FONT
    cell = sheet.cell(row=subtotal_row, column=6, value=delivery_total)
    cell.font = _SUBTOTAL_FONT
    cell.number_format = _CURRENCY_FMT
    cell.border = _TOP_BORDER

    sheet.freeze_panes = "A2"


def _populate_summary_sheet(
    sheet,
    *,
    company_name: str,
    project_name: str,
    params: ScenarioParams,
    project_duration_days: int,
    breakdown: QuoteBreakdown,
) -> None:
    sheet.title = "Quote Summary"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 18

    sheet.cell(row=1, column=1, value=company_name).font = _TITLE_FONT
    sheet.cell(row=2, column=1, value="TRAFFIC CONTROL COST ESTIMATE").font = _SUBTITLE_FONT
    sheet.cell(row=3, column=1, value=datetime.now().strftime("%Y-%m-%d"))

    sheet.cell(row=5, column=1, value=f"Project: {project_name}").font = _SUBTOTAL_FONT
    # UX-11: display name, not the raw "lane" / "shoulder" enum.
    closure = scenario_display_name(params)
    road = params.road_type.replace("_", " ")
    sheet.cell(
        row=6,
        column=1,
        value=(f"Speed {params.speed_mph} mph  |  Closure: {closure}  |  Road: {road}"),
    )
    sheet.cell(
        row=7,
        column=1,
        value=(
            f"Work zone: {params.work_zone_length_ft:.0f} ft  |  "
            f"Duration: {project_duration_days} day(s)"
            + ("  |  Night operation" if params.is_night else "")
        ),
    )

    table_start = 10
    sheet.cell(row=table_start, column=1, value="Category").font = _HEADER_FONT
    sheet.cell(row=table_start, column=2, value="Amount").font = _HEADER_FONT
    sheet.cell(row=table_start, column=1).fill = _HEADER_FILL
    sheet.cell(row=table_start, column=2).fill = _HEADER_FILL
    sheet.cell(row=table_start, column=2).alignment = Alignment(horizontal="right")

    rows: list[tuple[str, float, bool]] = [
        ("Equipment Rental", breakdown.equipment_total, False),
        ("Labor", breakdown.labor_total, False),
        ("Delivery & Logistics", breakdown.delivery_total, False),
        ("SUBTOTAL", breakdown.subtotal, True),
        (f"Overhead ({breakdown.overhead_pct * 100:.0f}%)", breakdown.overhead, False),
        (f"Profit ({breakdown.profit_pct * 100:.0f}%)", breakdown.profit, False),
        ("TOTAL ESTIMATE", breakdown.total, True),
    ]
    for offset, (label, amount, emphasize) in enumerate(rows, start=1):
        row_idx = table_start + offset
        label_cell = sheet.cell(row=row_idx, column=1, value=label)
        amount_cell = sheet.cell(row=row_idx, column=2, value=amount)
        amount_cell.number_format = _CURRENCY_FMT
        amount_cell.alignment = Alignment(horizontal="right")
        if emphasize:
            label_cell.font = _SUBTOTAL_FONT
            amount_cell.font = _SUBTOTAL_FONT
            label_cell.border = _TOP_BORDER
            amount_cell.border = _TOP_BORDER
        if label == "TOTAL ESTIMATE":
            label_cell.font = _TOTAL_FONT
            amount_cell.font = _TOTAL_FONT


def _populate_terms_sheet(sheet) -> None:
    sheet.title = "Terms & Conditions"
    sheet.sheet_view.showGridLines = False
    for col_letter, width in {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18}.items():
        sheet.column_dimensions[col_letter].width = width

    sheet.cell(row=1, column=1, value="TERMS AND CONDITIONS").font = _SUBTITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    current_row = 3
    for term in _TERMS:
        sheet.cell(row=current_row, column=1, value=term).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        sheet.row_dimensions[current_row].height = 24
        current_row += 1

    current_row += 1
    footer = sheet.cell(
        row=current_row,
        column=1,
        value="Generated by Conestruct — conestruct.com",
    )
    footer.font = _FOOTNOTE_FONT
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_quote(
    placements: list[DevicePlacement],
    params: ScenarioParams,
    output_path: str = "quote.xlsx",
    company_name: str = "Colorado Barricade Co.",
    project_name: str = "Untitled Project",
    project_duration_days: int = 1,
    num_flaggers: int = 0,
    tcs_hours_per_day: float = 2.0,
    num_setup_crew: int = 2,
    setup_hours: float = 1.5,
    takedown_hours: float = 1.0,
    delivery_trips: int = 2,
    delivery_distance_miles: float = 20.0,
    flagger_hourly_rate: float = 55.0,
    tcs_hourly_rate: float = 75.0,
    crew_hourly_rate: float = 45.0,
    mileage_rate: float = 3.50,
    min_trip_charge: float = 150.0,
    overhead_pct: float = 0.10,
    profit_pct: float = 0.10,
    night_multiplier: float = 1.5,
    applied_deltas: list[dict[str, Any]] | None = None,
) -> tuple[str, QuoteBreakdown]:
    """Write a contractor pricing workbook for ``placements``.

    ``applied_deltas`` are the fired jurisdiction count deltas (issue
    #151/#154); the quote joins the shared aggregation path so a fired delta
    is priced on the same rows the XLSX bid document and on-sheet summary
    render.  Omit them (or pass ``None``) for a plain, delta-free quote —
    identical to the pre-#154 output.

    Returns ``(output_path, breakdown)`` where ``breakdown`` is a structured
    ``QuoteBreakdown`` with all line items, subtotals, markup, and the final
    total — ready to drive a UI drill-down or further programmatic use.
    """
    rows = aggregate_device_rows_with_deltas(placements, applied_deltas)

    equipment_lines = _compute_equipment_lines(rows, params, project_duration_days)
    labor_lines = _compute_labor_lines(
        is_night=params.is_night,
        project_duration_days=project_duration_days,
        num_flaggers=num_flaggers,
        tcs_hours_per_day=tcs_hours_per_day,
        num_setup_crew=num_setup_crew,
        setup_hours=setup_hours,
        takedown_hours=takedown_hours,
        flagger_hourly_rate=flagger_hourly_rate,
        tcs_hourly_rate=tcs_hourly_rate,
        crew_hourly_rate=crew_hourly_rate,
        night_multiplier=night_multiplier,
    )
    delivery_lines = _compute_delivery_lines(
        delivery_trips=delivery_trips,
        delivery_distance_miles=delivery_distance_miles,
        mileage_rate=mileage_rate,
        min_trip_charge=min_trip_charge,
    )

    # #135 ruling (Refs #185): every line.extended above is already a
    # cent value; the sums below therefore combine rounded values, and the
    # markup products are themselves decided at cent resolution.  _cents on
    # the sums re-normalizes binary float residue (0.01-step values are not
    # exactly representable), never changes the decimal amount.
    equipment_total = _cents(sum(line.extended for line in equipment_lines))
    labor_total = _cents(sum(line.extended for line in labor_lines))
    delivery_total = _cents(sum(line.extended for line in delivery_lines))
    subtotal = _cents(equipment_total + labor_total + delivery_total)
    overhead = _cents(subtotal * overhead_pct)
    profit = _cents((subtotal + overhead) * profit_pct)
    total = _cents(subtotal + overhead + profit)

    breakdown = QuoteBreakdown(
        equipment_lines=equipment_lines,
        labor_lines=labor_lines,
        delivery_lines=delivery_lines,
        is_night=params.is_night,
        night_multiplier=night_multiplier,
        overhead_pct=overhead_pct,
        profit_pct=profit_pct,
        equipment_total=equipment_total,
        labor_total=labor_total,
        delivery_total=delivery_total,
        subtotal=subtotal,
        overhead=overhead,
        profit=profit,
        total=total,
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    equipment_sheet = workbook.create_sheet("Equipment Detail")
    labor_sheet = workbook.create_sheet("Labor Detail")
    delivery_sheet = workbook.create_sheet("Delivery & Logistics")
    terms_sheet = workbook.create_sheet("Terms & Conditions")

    _populate_summary_sheet(
        summary_sheet,
        company_name=company_name,
        project_name=project_name,
        params=params,
        project_duration_days=project_duration_days,
        breakdown=breakdown,
    )
    _populate_equipment_sheet(equipment_sheet, equipment_lines)
    _populate_labor_sheet(
        labor_sheet,
        labor_lines,
        is_night=params.is_night,
        night_multiplier=night_multiplier,
    )
    _populate_delivery_sheet(delivery_sheet, delivery_lines)
    _populate_terms_sheet(terms_sheet)

    for sheet in workbook.worksheets:
        _apply_left_align(sheet)

    workbook.save(output_path)
    return output_path, breakdown


# ---------------------------------------------------------------------------
# Smoke test
# ---------------------------------------------------------------------------


if __name__ == "__main__":
    import os

    from src.generation.layout import generate_shoulder_closure_divided

    smoke_params = ScenarioParams(
        speed_mph=55,
        num_lanes=2,
        closure_type="shoulder",
        road_type="freeway",
        work_zone_length_ft=800.0,
        lane_width_ft=12.0,
        is_divided=True,
        jurisdiction="CDOT",
    )
    smoke_placements = generate_shoulder_closure_divided(smoke_params)

    duration = 3
    flaggers = 2
    distance = 25.0

    path, breakdown = generate_quote(
        smoke_placements,
        smoke_params,
        output_path="quote.xlsx",
        project_duration_days=duration,
        num_flaggers=flaggers,
        delivery_distance_miles=distance,
    )

    size_bytes = os.path.getsize(path)
    print(f"Wrote {path} ({size_bytes:,} bytes)")
    print(
        f"Placements: {len(smoke_placements)}  "
        f"Equipment lines: {len(breakdown.equipment_lines)}  "
        f"Labor lines: {len(breakdown.labor_lines)}"
    )
    print()
    print("Cost breakdown")
    print("-" * 44)
    print(f"  Equipment Rental     ${breakdown.equipment_total:>12,.2f}")
    print(f"  Labor                ${breakdown.labor_total:>12,.2f}")
    print(f"  Delivery & Logistics ${breakdown.delivery_total:>12,.2f}")
    print("-" * 44)
    print(f"  SUBTOTAL             ${breakdown.subtotal:>12,.2f}")
    print(f"  Overhead ({breakdown.overhead_pct * 100:.0f}%)       ${breakdown.overhead:>12,.2f}")
    print(f"  Profit ({breakdown.profit_pct * 100:.0f}%)         ${breakdown.profit:>12,.2f}")
    print("-" * 44)
    print(f"  TOTAL ESTIMATE       ${breakdown.total:>12,.2f}")
    print()

    print("Equipment lines:")
    for line in breakdown.equipment_lines:
        label_part = f" {line.label}" if line.label else ""
        print(
            f"  [{line.item_number:>2}] {line.device_type}{label_part:<14}  "
            f"qty {line.qty:>3} x ${line.daily_rate:>6.2f} x {line.days}d  "
            f"= ${line.extended:>9,.2f}"
        )
    print()
    print("Labor lines:")
    for line in breakdown.labor_lines:
        print(
            f"  {line.role:<18}  {line.personnel} x {line.hours_per_day:>4.1f} hr x "
            f"{line.days}d x ${line.rate:>5.2f}/hr = ${line.extended:>9,.2f}"
        )
    print()
    print("Delivery lines:")
    for line in breakdown.delivery_lines:
        print(
            f"  {line.item:<22}  {line.trips} trips x max({line.distance_miles} mi x "
            f"${line.rate_per_mile}/mi, ${line.min_trip_charge}) = ${line.extended:>9,.2f}"
        )
