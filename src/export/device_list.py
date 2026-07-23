"""Build a CDOT Spec 630 device-list spreadsheet from a placement list.

Aggregates the per-device placements emitted by the layout engine into
the by-pay-item summary that appears on the engineer's plan set as the
"Tabulation of Quantities" page.

Authoritative sources:
  - CDOT Standard Specifications Section 630 (pay item names, units)
  - CDOT M&S Standard Plan S-630-1 (typical device-list format)
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.rules.device_aggregation import AggregatedDeviceRow, aggregate_device_rows
from src.rules.devices import DEVICE_CATALOG, DeviceType, cone_display_name
from src.rules.jurisdiction import aggregate_device_rows_with_deltas
from src.rules.sign_codes import substitute_sign_description
from src.rules.validators import DevicePlacement, ScenarioParams, scenario_display_name

# Light-gray header fill from the V1 spec.
_HEADER_FILL: PatternFill = PatternFill(
    fill_type="solid", start_color="FFD9D9D9", end_color="FFD9D9D9"
)
_HEADER_FONT: Font = Font(bold=True)
_LEFT_ALIGN: Alignment = Alignment(horizontal="left", vertical="center")


def _apply_left_align(sheet) -> None:
    """Left-align every populated cell on ``sheet``.

    Excel right-aligns numbers and left-aligns text by default; this
    forces a consistent visual alignment across columns.  Bold/fill
    on header cells is preserved (alignment is independent of font).
    """
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.alignment = _LEFT_ALIGN


_COLUMN_WIDTHS: dict[str, int] = {
    "A": 8,  # Item #
    "B": 25,  # Device Type
    "C": 40,  # Description
    "D": 25,  # CDOT Pay Item
    "E": 8,  # Unit
    "F": 10,  # Quantity
    "G": 50,  # Notes
}

_DEVICE_LIST_HEADERS: tuple[str, ...] = (
    "Item #",
    "Device Type",
    "Description",
    "CDOT Pay Item #",
    "Unit",
    "Quantity",
    "Notes",
)

_SIGN_GENERIC_NOTE: str = (
    "Unit is EACH for V1; CDOT Spec 630 bills by SF — convert when sign sizes are known."
)
_BARRICADE_TYPE_II_NOTE: str = (
    "Unit is EACH for V1; CDOT Spec 630 bills Type I/II under Construction "
    "Traffic Sign (Special) by SF — convert when panel dimensions are known."
)
_CHANNELIZER_OPTIONAL_NOTE: str = (
    "Optional — apply probability weight as appropriate per engineer discretion."
)
# Note for a fired jurisdiction count-delta whose device id has no single
# CDOT pay item (e.g. ``advance_warning_signs`` — a set of W-series signs
# billed per SF, issue #151).  Follows the shipped ``_SIGN_GENERIC_NOTE``
# convention and reuses its "CDOT Spec 630 ... by SF" citation — no new
# citation, no fabricated pay item (rule 10 / 2026-07-22 ruling).  Any
# future unmapped device that fires must either join ``_DELTA_DEVICE_TYPE``
# or have this wording checked against its own billing basis.
_JURISDICTION_UNMAPPED_NOTE: str = (
    "Jurisdiction-required — {doc}. Unit is EACH for V1; CDOT Spec 630 bills "
    "these as individual W-series signs by SF — itemize per sign type; no "
    "single pay item."
)
_JURISDICTION_MAPPED_NOTE: str = "Jurisdiction-required — {doc}."


def _jurisdiction_doc(source: dict[str, Any] | None) -> str:
    """Human doc name from a delta ``source`` block, for the Notes column."""
    return (source or {}).get("doc", "jurisdiction requirement")


def _row_for(
    item_number: int,
    row: AggregatedDeviceRow,
    params: ScenarioParams,
) -> tuple[int, str, str, str, str, int, str]:
    """Build a single Device-List row tuple in column order.

    ``row.label`` is the aggregation key (schedule key for signs);
    ``row.representative`` is one placement from the group (lowest station)
    so the shared substitution helper can resolve station-dependent
    parametric values — keeping the XLSX descriptions identical to the
    PDF schedule / off-page table.

    Jurisdiction count-delta rows (issue #151): a delta-only add with no
    backing placement (``display_override`` set) renders as a bid line
    from its mapped ``DeviceType`` — a real CDOT pay item + unit — or, for
    an unmapped device id (no single pay item, e.g. advance warning
    signs), as an honest pay-item-less line rather than a fabricated one
    (rule 10).  A topped-up real row renders normally and gains a
    jurisdiction-required note.
    """
    device_type = row.device_type
    quantity = row.quantity

    # Delta-only add (no backing placement).
    if row.display_override is not None:
        doc = _jurisdiction_doc(row.jurisdiction_source)
        if device_type is not None:
            spec = DEVICE_CATALOG[device_type]
            return (
                item_number,
                row.display_override,
                spec.description,
                spec.cdot_pay_item_number or "TODO",
                spec.unit,
                quantity,
                _JURISDICTION_MAPPED_NOTE.format(doc=doc),
            )
        # Unmapped device id: honest row, no fabricated pay item.
        return (
            item_number,
            f"{row.display_override} (jurisdiction-required)",
            row.display_override,
            "—",
            "EACH",
            quantity,
            _JURISDICTION_UNMAPPED_NOTE.format(doc=doc),
        )

    label = row.label
    representative = row.representative
    spec = DEVICE_CATALOG[device_type]
    pay_item_number = spec.cdot_pay_item_number or "TODO"

    if device_type == DeviceType.SIGN_GENERIC:
        if label is None:
            description = "Generic construction sign (unlabeled)"
            type_label = "SIGN_GENERIC (unlabeled)"
        else:
            station_ft = representative.station_ft if representative is not None else 0.0
            code, human = substitute_sign_description(label, station_ft, params)
            description = f"{code} {human}".strip() if human != code else code
            type_label = "SIGN_GENERIC"
        unit = "EACH"  # V1 override; catalog says SF.
        notes = _SIGN_GENERIC_NOTE
    elif device_type == DeviceType.BARRICADE_TYPE_II:
        description = spec.description
        type_label = device_type.value
        unit = "EACH"  # V1 override; catalog says SF per Table 630-7 footnote.
        notes = _BARRICADE_TYPE_II_NOTE
    elif device_type == DeviceType.CONE:
        # Resolve the size for the posted speed (§6F.65) — same helper the
        # narrative, UI breakdown, and plan-sheet legend use (Refs #101).
        description = cone_display_name(params.speed_mph)
        type_label = device_type.value
        unit = spec.unit
        notes = ""
    else:
        description = spec.description
        type_label = device_type.value
        unit = spec.unit
        notes = _CHANNELIZER_OPTIONAL_NOTE if device_type == DeviceType.CHANNELIZER_OPTIONAL else ""

    # A real row a delta topped up (issue #151): keep its catalog identity,
    # append the jurisdiction provenance to whatever note it already has.
    if row.jurisdiction_required:
        jur_note = _JURISDICTION_MAPPED_NOTE.format(doc=_jurisdiction_doc(row.jurisdiction_source))
        notes = f"{notes} {jur_note}".strip() if notes else jur_note

    return (item_number, type_label, description, pay_item_number, unit, quantity, notes)


def _populate_device_list_sheet(
    sheet,
    placements: list[DevicePlacement],
    params: ScenarioParams,
    applied_deltas: list[dict[str, Any]] | None = None,
) -> list[AggregatedDeviceRow]:
    """Write the Device-List sheet and return the aggregated rows."""
    sheet.title = "Device List"
    sheet.append(_DEVICE_LIST_HEADERS)
    for col_letter in _COLUMN_WIDTHS:
        sheet.column_dimensions[col_letter].width = _COLUMN_WIDTHS[col_letter]

    header_row = sheet[1]
    for cell in header_row:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    # Aggregation + ordering + fired jurisdiction count deltas live in the
    # shared helper (issue #150/#151) so the screen breakdown, this bid
    # document, and the on-sheet summary carry identical quantities.
    aggregated = aggregate_device_rows_with_deltas(placements, applied_deltas)

    for item_number, row in enumerate(aggregated, start=1):
        sheet.append(_row_for(item_number, row, params))
        sheet.cell(row=item_number + 1, column=6).number_format = "0"

    sheet.freeze_panes = "A2"
    return aggregated


def _populate_summary_sheet(
    sheet,
    params: ScenarioParams,
    aggregated_rows: list[AggregatedDeviceRow],
) -> None:
    """Write the Summary sheet."""
    sheet.title = "Summary"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 30

    # Sum of the aggregated (delta-aware) row quantities, not len(placements):
    # a fired jurisdiction count-delta adds a required device with no backing
    # placement, so the bid total must count the rows, not the raw layout
    # (issue #151).  Equals len(placements) whenever no delta fires.
    rows = (
        ("Total device count", sum(r.quantity for r in aggregated_rows)),
        ("Total unique device types", len(aggregated_rows)),
        ("Speed (mph)", params.speed_mph),
        # UX-11: display name, not the raw "lane" / "shoulder" enum.
        ("Closure type", scenario_display_name(params)),
        ("Work zone length (ft)", params.work_zone_length_ft),
        ("Jurisdiction", params.jurisdiction),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )

    sheet.append(("Field", "Value"))
    header_row = sheet[1]
    for cell in header_row:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for label, value in rows:
        sheet.append((label, value))


def export_device_list(
    placements: list[DevicePlacement],
    params: ScenarioParams,
    output_path: str = "device_list.xlsx",
    applied_deltas: list[dict[str, Any]] | None = None,
) -> str:
    """Write a CDOT-format device-list workbook for ``placements``.

    ``applied_deltas`` are the fired jurisdiction count deltas (issue
    #151); a jurisdiction-required device with no backing placement is
    added to the device list and the totals, so the bid document matches
    the on-screen breakdown.  Omit them (or pass ``None``) for a plain,
    delta-free workbook.

    Returns the absolute or relative path written, matching ``output_path``.
    """
    workbook = Workbook()
    device_sheet = workbook.active
    aggregated = _populate_device_list_sheet(device_sheet, placements, params, applied_deltas)
    summary_sheet = workbook.create_sheet("Summary")
    _populate_summary_sheet(summary_sheet, params, aggregated)
    for sheet in workbook.worksheets:
        _apply_left_align(sheet)
    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    import os

    from src.generation.layout import generate_shoulder_closure_divided

    params = ScenarioParams(
        speed_mph=55,
        num_lanes=2,
        closure_type="shoulder",
        road_type="freeway",
        work_zone_length_ft=800.0,
        lane_width_ft=12.0,
        is_divided=True,
        jurisdiction="CDOT",
    )
    placements = generate_shoulder_closure_divided(params)
    output = export_device_list(placements, params, "device_list.xlsx")

    size_bytes = os.path.getsize(output)
    print(f"Wrote {output} ({size_bytes} bytes)")
    print()

    aggregated = aggregate_device_rows(placements)
    print(f"{'#':>3}  {'Device Type':25s}  {'Label':12s}  {'Qty':>4s}")
    print("-" * 52)
    for i, row in enumerate(aggregated, start=1):
        print(f"{i:>3}  {row.device_type.value:25s}  {(row.label or '-'):12s}  {row.quantity:>4d}")
    print("-" * 52)
    print(f"     {'TOTAL':25s}  {'':12s}  {len(placements):>4d}")
