"""Dev-only replication-snapshot markdown builder (Refs #102).

TEMPORARY DIAGNOSTIC SCAFFOLDING — not a V1 product feature. Delete this
file together with the ``/render/replication-snapshot`` endpoint in
``render_api.py`` and the frontend ``DebugSnapshotButton`` when no longer
needed.

Produces sections 3-6 of the replication snapshot (the backend halves);
the frontend button prepends sections 1-2 (location / road detection and
the scenario default-vs-changed diff, which only the frontend can know).

Rule #3 (single source of truth) applies to this tool too: every section
below is sourced by calling the SAME shared function the corresponding
real deliverable uses — ``_audit_projection_for`` (audit endpoints),
``_run_quote`` (quote endpoints), ``export_device_list`` (the shipped
XLSX, read back verbatim), and ``render_crew_narrative_markdown`` (the
crew narrative's single content path). Nothing here re-derives or
reformats generation logic; drift between this dump and the real
deliverables is impossible by construction.

The plan-sheet PDF's visual rendering is deliberately excluded (spec):
its textual content (TA reference, case, parameters, sign schedule) is
already present in the audit projection and device list below.
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

from fastapi import HTTPException

if TYPE_CHECKING:  # imported for type hints only — avoids the runtime cycle
    from src.api.render_api import QuoteRequest


def _cell(value: Any) -> str:
    """One markdown table cell: pipe-safe, newline-free, empty-safe."""
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def _markdown_table(headers: tuple[str, ...], rows: list[tuple]) -> str:
    lines = [
        "| " + " | ".join(_cell(h) for h in headers) + " |",
        "|" + "---|" * len(headers),
    ]
    for row in rows:
        lines.append("| " + " | ".join(_cell(v) for v in row) + " |")
    return "\n".join(lines)


def build_snapshot_markdown(req: QuoteRequest) -> str:
    """Assemble the backend sections of the replication snapshot.

    Imports from ``render_api`` are deferred to call time: ``render_api``
    imports this module for its endpoint, so a top-level import here
    would be circular.
    """
    from openpyxl import load_workbook

    from src.api.render_api import _audit_projection_for, _placements_for, _run_quote
    from src.export.device_list import export_device_list
    from src.narrative.crew_narrative import render_crew_narrative_markdown

    scenario = req.scenario

    # Baked into the Modal image at deploy time (modal_app.py ._git_sha);
    # "unknown" when unstamped — same honest sentinel as /healthz.
    git_sha = os.environ.get("GIT_SHA", "unknown")

    # The shared validated pipeline — the same gates as every deliverable.
    # A 400 refusal (eligibility gates, geometry validation) is a state
    # the snapshot exists to document, not an error: emit a REFUSED
    # section quoting the detail verbatim instead of dying (issue #178).
    # Anything else propagates unchanged.
    try:
        placements, params, site_adj, night_adj, approaches = _placements_for(scenario)
    except HTTPException as exc:
        if exc.status_code != 400:
            raise
        detail = exc.detail
        detail_text = (
            json.dumps(detail, indent=2, sort_keys=True, ensure_ascii=False)
            if isinstance(detail, (dict, list))
            else str(detail)
        )
        return "\n".join(
            [
                "<!-- Backend sections of the replication snapshot (Refs #102)."
                " Generation was refused; see below. -->",
                "",
                "## 3 GENERATION REFUSED",
                "",
                f"Backend: GIT_SHA={git_sha}",
                "",
                "The generation pipeline refused this scenario (HTTP 400).",
                "The refusal detail, verbatim:",
                "",
                "```json",
                detail_text,
                "```",
                "",
                "Sections 4-6 (quote, device list, crew narrative) are"
                " underivable: every backend section derives from the same"
                " pipeline that refused; nothing here re-computes what the"
                " deliverables would not produce.",
            ]
        )

    # Section 3 — the complete audit projection, verbatim.
    projection = _audit_projection_for(scenario)
    projection_json = json.dumps(projection, indent=2, sort_keys=True, ensure_ascii=False)

    # Section 4 — the same QuoteBreakdown the XLSX writer builds from.
    _, breakdown = _run_quote(req)
    equipment_table = _markdown_table(
        (
            "Item",
            "Device type",
            "Label",
            "Description",
            "Qty",
            "Unit",
            "Daily rate",
            "Days",
            "Extended",
            "Note",
        ),
        [
            (
                line.item_number,
                line.device_type,
                line.label,
                line.description,
                line.qty,
                line.unit,
                f"{line.daily_rate:.2f}",
                line.days,
                f"{line.extended:.2f}",
                line.note,
            )
            for line in breakdown.equipment_lines
        ],
    )
    labor_table = _markdown_table(
        ("Role", "Personnel", "Hours/day", "Days", "Rate", "Extended"),
        [
            (
                line.role,
                line.personnel,
                line.hours_per_day,
                line.days,
                f"{line.rate:.2f}",
                f"{line.extended:.2f}",
            )
            for line in breakdown.labor_lines
        ],
    )
    delivery_table = _markdown_table(
        ("Item", "Trips", "Distance (mi)", "Rate/mi", "Min charge", "Extended"),
        [
            (
                line.item,
                line.trips,
                line.distance_miles,
                f"{line.rate_per_mile:.2f}",
                f"{line.min_trip_charge:.2f}",
                f"{line.extended:.2f}",
            )
            for line in breakdown.delivery_lines
        ],
    )
    totals = (
        f"- Night operation: {breakdown.is_night}"
        f" (labor multiplier {breakdown.night_multiplier:.1f}x)\n"
    ) + "\n".join(
        f"- {label}: {value:.2f}"
        for label, value in (
            ("Equipment total", breakdown.equipment_total),
            ("Labor total", breakdown.labor_total),
            ("Delivery total", breakdown.delivery_total),
            ("Subtotal", breakdown.subtotal),
            (f"Overhead ({breakdown.overhead_pct * 100:.0f}%)", breakdown.overhead),
            (f"Profit ({breakdown.profit_pct * 100:.0f}%)", breakdown.profit),
            ("TOTAL", breakdown.total),
        )
    )

    # Section 5 — the shipped Device List workbook, read back verbatim.
    fd, raw_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    xlsx_path = Path(raw_path)
    try:
        export_device_list(placements, params, output_path=str(xlsx_path))
        workbook = load_workbook(str(xlsx_path), read_only=True)
        try:
            sheet_rows = list(workbook["Device List"].iter_rows(values_only=True))
        finally:
            # read_only keeps the file handle open — close before the
            # unlink below or a raise here turns into PermissionError on
            # Windows, masking the real error.
            workbook.close()
    finally:
        xlsx_path.unlink(missing_ok=True)
    device_table = _markdown_table(tuple(sheet_rows[0]), sheet_rows[1:])

    # Section 6 — the crew narrative's single content path (no file I/O).
    narrative_md = render_crew_narrative_markdown(
        placements,
        params,
        site_adjustments=site_adj,
        night_adjustments=night_adj,
        pilot_car=getattr(scenario, "pilotCar", False),
        # Same ApproachParams the generator got (near_intersection only);
        # the builder raises rather than emit a partial narrative (#117).
        approaches=approaches,
    )

    return "\n".join(
        [
            "<!-- Backend sections of the replication snapshot (Refs #102)."
            " The plan-sheet PDF's visual rendering is deliberately excluded;"
            " its textual content appears in the audit projection below. -->",
            "",
            "## 3 Audit projection (backend, verbatim)",
            "",
            f"Backend: GIT_SHA={git_sha}",
            "",
            "```json",
            projection_json,
            "```",
            "",
            "## 4 Quote (backend, from the same QuoteBreakdown the XLSX ships)",
            "",
            "### Equipment Detail",
            "",
            equipment_table,
            "",
            "### Labor Detail",
            "",
            labor_table,
            "",
            "### Delivery & Logistics",
            "",
            delivery_table,
            "",
            "### Totals",
            "",
            totals,
            "",
            "## 5 Device list (backend, the shipped XLSX read back)",
            "",
            device_table,
            "",
            "## 6 Crew narrative (backend, verbatim markdown)",
            "",
            narrative_md,
        ]
    )
